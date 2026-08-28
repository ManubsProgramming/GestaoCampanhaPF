from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from rest_framework.views import APIView

from usuarios.permissions import EhAdministrador

from .services import filtrar_pessoas


class ExportarExcelView(APIView):
    permission_classes = [EhAdministrador]

    def get(self, request):
        pessoas = filtrar_pessoas(request)

        workbook = Workbook()

        planilha = workbook.active
        planilha.title = "Pessoas cadastradas"

        cabecalhos = [
            "Nome",
            "CPF",
            "Data de nascimento",
            "Título de eleitor",
            "Telefone",
            "Região",
            "Localidade",
            "Tipo da localidade",
            "Rua",
            "Número",
            "Complemento",
            "Cadastrado por",
            "Status",
            "Data do cadastro",
        ]

        planilha.append(cabecalhos)

        for celula in planilha[1]:
            celula.font = Font(
                bold=True
            )

            celula.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        for pessoa in pessoas:
            cadastrador = (
                pessoa.cadastrada_por.get_full_name()
                or pessoa.cadastrada_por.username
            )

            nascimento = ""

            if pessoa.data_nascimento:
                nascimento = (
                    pessoa.data_nascimento.strftime(
                        "%d/%m/%Y"
                    )
                )

            planilha.append([
                pessoa.nome_completo,
                pessoa.cpf,
                nascimento,
                pessoa.titulo_eleitor,
                pessoa.telefone,
                pessoa.regiao.nome,
                pessoa.localidade.nome,
                pessoa.localidade.get_tipo_display(),
                pessoa.rua.nome,
                pessoa.numero,
                pessoa.complemento,
                cadastrador,
                pessoa.get_status_display(),
                pessoa.criado_em.strftime(
                    "%d/%m/%Y %H:%M"
                ),
            ])

        larguras = {
            1: 30,
            2: 16,
            3: 18,
            4: 18,
            5: 18,
            6: 25,
            7: 25,
            8: 20,
            9: 30,
            10: 10,
            11: 25,
            12: 25,
            13: 12,
            14: 20,
        }

        for indice, largura in larguras.items():
            coluna = get_column_letter(
                indice
            )

            planilha.column_dimensions[
                coluna
            ].width = largura

        planilha.freeze_panes = "A2"

        response = HttpResponse(
            content_type=(
                "application/"
                "vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
        )

        response["Content-Disposition"] = (
            'attachment; '
            'filename="relatorio_pessoas_cadastradas.xlsx"'
        )

        workbook.save(response)

        return response